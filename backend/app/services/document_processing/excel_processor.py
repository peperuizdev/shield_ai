"""
Excel Document Processor for Shield AI

Specialized processor for Excel (.xlsx) files using openpyxl for data extraction.
Handles multiple worksheets, cells, and converts tabular data to plain text.
"""

import io
import logging
from typing import Dict, List, Any
from zipfile import BadZipFile

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base import DocumentProcessor, DocumentProcessingError, DocumentValidationError

logger = logging.getLogger(__name__)


class ExcelProcessor(DocumentProcessor):
    """
    Excel document processor using openpyxl for data extraction.
    
    Implements the DocumentProcessor interface for .xlsx files,
    providing structured data extraction from multiple worksheets.
    """
    
    def __init__(self):
        """Initialize Excel processor."""
        super().__init__()
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel processing. Install with: pip install openpyxl==3.1.2")
    
    def extract_text(self, file_content: bytes) -> str:
        """
        Extract text from Excel workbook using openpyxl.
        
        Args:
            file_content (bytes): Raw .xlsx file bytes
            
        Returns:
            str: Extracted plain text from all worksheets
            
        Raises:
            DocumentProcessingError: If text extraction fails
        """
        try:
            self.logger.info("Extracting text from Excel workbook")
            
            xlsx_file = io.BytesIO(file_content)
            workbook = load_workbook(xlsx_file, data_only=True)
            
            extracted_text = ""
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                self.logger.debug(f"Processing worksheet: {sheet_name}")
                extracted_text += f"--- Hoja: {sheet_name} ---\n"
                
                # Get the used range of the worksheet
                if worksheet.max_row == 1 and worksheet.max_column == 1:
                    # Check if the single cell has content
                    cell_value = worksheet.cell(1, 1).value
                    if not cell_value:
                        extracted_text += "Hoja vacía\n\n"
                        continue
                
                # Extract data row by row
                for row_num in range(1, worksheet.max_row + 1):
                    row_data = []
                    
                    for col_num in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row_num, col_num)
                        cell_value = cell.value
                        
                        if cell_value is not None:
                            # Convert cell value to string
                            if isinstance(cell_value, (int, float)):
                                # Handle numbers (including dates stored as numbers)
                                if cell.number_format and 'd' in cell.number_format.lower():
                                    # Likely a date format
                                    row_data.append(str(cell_value))
                                else:
                                    row_data.append(str(cell_value))
                            else:
                                row_data.append(str(cell_value).strip())
                    
                    # Add row to text if it has content
                    if row_data and any(data.strip() for data in row_data if data):
                        row_text = " | ".join([data for data in row_data if data.strip()])
                        if row_text.strip():
                            extracted_text += row_text + "\n"
                
                extracted_text += "\n"
            
            if not extracted_text.strip() or extracted_text.strip() == "":
                raise DocumentProcessingError("No data could be extracted from Excel workbook")
            
            self.logger.info("Excel text extraction completed successfully")
            return extracted_text.strip()
            
        except Exception as e:
            self.logger.error(f"Excel text extraction failed: {str(e)}")
            raise DocumentProcessingError(f"Failed to extract text from Excel workbook: {str(e)}")
    
    def validate_file(self, file_content: bytes, filename: str) -> bool:
        """
        Validate Excel file structure and readability.
        
        Args:
            file_content (bytes): Raw file bytes
            filename (str): Original filename
            
        Returns:
            bool: True if file is valid .xlsx
            
        Raises:
            DocumentValidationError: If validation fails
        """
        try:
            if len(file_content) == 0:
                raise DocumentValidationError("Excel file is empty")
            
            # Check if it's a .xlsx file (ZIP-based format)
            if not file_content.startswith(b'PK'):
                raise DocumentValidationError("File is not a valid .xlsx format (legacy .xls files not supported)")
            
            xlsx_file = io.BytesIO(file_content)
            
            try:
                workbook = load_workbook(xlsx_file, data_only=True)
                
                # Check if workbook has sheets
                if len(workbook.sheetnames) == 0:
                    raise DocumentValidationError("Excel workbook contains no worksheets")
                
                self.logger.debug(f"Excel workbook has {len(workbook.sheetnames)} worksheets")
                
                # Test if we can access at least one worksheet
                first_sheet = workbook[workbook.sheetnames[0]]
                
                # Check if workbook has any data
                has_content = False
                for sheet_name in workbook.sheetnames:
                    worksheet = workbook[sheet_name]
                    
                    # Check if sheet has data beyond empty cells
                    if worksheet.max_row > 1 or worksheet.max_column > 1:
                        has_content = True
                        break
                    else:
                        # Check the single cell case
                        cell_value = worksheet.cell(1, 1).value
                        if cell_value is not None and str(cell_value).strip():
                            has_content = True
                            break
                
                if not has_content:
                    self.logger.warning(f"Excel workbook {filename} appears to contain no data")
            
            except InvalidFileException:
                raise DocumentValidationError("Invalid .xlsx file structure")
            except BadZipFile:
                raise DocumentValidationError("Corrupted .xlsx file (invalid ZIP structure)")
            
            self.logger.info(f"Excel validation successful: {filename}")
            return True
            
        except DocumentValidationError:
            raise
        except Exception as e:
            self.logger.error(f"Excel validation failed for {filename}: {str(e)}")
            raise DocumentValidationError(f"Excel validation failed: {str(e)}")
    
    def get_supported_extensions(self) -> List[str]:
        """
        Get list of Excel file extensions supported.
        
        Returns:
            List[str]: Supported extensions
        """
        return ['.xlsx', '.XLSX']
    
    def get_supported_mime_types(self) -> List[str]:
        """
        Get list of Excel MIME types supported.
        
        Returns:
            List[str]: Supported MIME types
        """
        return [
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ]
    
    def get_metadata(self, file_content: bytes) -> Dict[str, Any]:
        """
        Extract metadata from Excel workbook.
        
        Args:
            file_content (bytes): Raw file bytes
            
        Returns:
            Dict[str, Any]: Excel workbook metadata
        """
        try:
            xlsx_file = io.BytesIO(file_content)
            workbook = load_workbook(xlsx_file, data_only=True)
            
            metadata = {
                "file_size": len(file_content),
                "worksheet_count": len(workbook.sheetnames),
                "worksheet_names": workbook.sheetnames
            }
            
            # Extract workbook properties if available
            if hasattr(workbook, 'properties') and workbook.properties:
                props = workbook.properties
                
                if props.title:
                    metadata["title"] = props.title
                if props.creator:
                    metadata["creator"] = props.creator
                if props.description:
                    metadata["description"] = props.description
                if props.created:
                    metadata["created"] = props.created.isoformat() if props.created else None
                if props.modified:
                    metadata["modified"] = props.modified.isoformat() if props.modified else None
                if props.lastModifiedBy:
                    metadata["last_modified_by"] = props.lastModifiedBy
            
            # Calculate data statistics
            total_cells = 0
            total_rows = 0
            non_empty_cells = 0
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                sheet_rows = worksheet.max_row
                sheet_cols = worksheet.max_column
                
                total_rows += sheet_rows
                total_cells += sheet_rows * sheet_cols
                
                # Count non-empty cells
                for row in range(1, sheet_rows + 1):
                    for col in range(1, sheet_cols + 1):
                        cell_value = worksheet.cell(row, col).value
                        if cell_value is not None and str(cell_value).strip():
                            non_empty_cells += 1
            
            metadata.update({
                "total_rows": total_rows,
                "total_cells": total_cells,
                "non_empty_cells": non_empty_cells,
                "data_density": round(non_empty_cells / max(total_cells, 1) * 100, 2)
            })
            
            return metadata
            
        except Exception as e:
            self.logger.warning(f"Could not extract Excel metadata: {str(e)}")
            return {"file_size": len(file_content)}